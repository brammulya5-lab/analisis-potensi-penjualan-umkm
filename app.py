import io
from pathlib import Path

import joblib
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import xgboost as xgb
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# KONFIGURASI HALAMAN

st.set_page_config(
    page_title="Analisis Potensi Penjualan UMKM",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="auto"
)

# DESIGN TOKENS (TEMA TERANG PROFESIONAL)

COLOR_BG          = "#f5f6fb"
COLOR_SURFACE     = "#ffffff"
COLOR_BORDER      = "#e4e7f2"
COLOR_INK         = "#161b33"
COLOR_MUTED       = "#6b7188"
COLOR_PRIMARY     = "#3730d1"
COLOR_PRIMARY_DK  = "#2a2299"
COLOR_PRIMARY_BG  = "#eeedfd"
COLOR_HIGH        = "#0f9d68"
COLOR_MED         = "#d68a12"
COLOR_LOW         = "#d64545"
COLOR_HIGH_BG     = "#e4f6ee"
COLOR_MED_BG      = "#fdf1de"
COLOR_LOW_BG      = "#fbe9e9"
COLOR_INFO        = "#1d6fd6"

CHART_COLOR_MAP = {
    "Potensi Tinggi": COLOR_HIGH,
    "Potensi Sedang": COLOR_MED,
    "Potensi Rendah": COLOR_LOW,
}
PLOTLY_TEMPLATE = "simple_white"
PLOTLY_FONT = dict(family="Inter, Segoe UI, sans-serif", color=COLOR_INK)

def hex_no_hash(hex_color):
    """Ubah '#RRGGBB' menjadi 'RRGGBB' (format yang dipakai openpyxl)."""
    return hex_color.lstrip("#").upper()

# CSS KUSTOM

st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@600;700;800&family=Inter:wght@400;500;600;700&display=swap');

    html, body, [class*="css"]  {{ font-family: 'Inter', 'Segoe UI', sans-serif; }}

    .stApp {{
        background-color: {COLOR_BG};
        color: {COLOR_INK};
    }}

    h1, h2, h3, h4 {{ font-family: 'Poppins', sans-serif; letter-spacing: -0.01em; }}

    .app-topbar {{
        background: linear-gradient(120deg, {COLOR_PRIMARY_DK} 0%, {COLOR_PRIMARY} 55%, #5b52e0 100%);
        border-radius: 18px;
        padding: 30px 34px;
        margin-bottom: 22px;
        box-shadow: 0 10px 24px -8px rgba(55, 48, 209, 0.35);
        position: relative;
        overflow: hidden;
    }}
    .app-topbar h1 {{
        color: #ffffff; font-size: 1.65rem; font-weight: 800; margin: 0 0 8px 0; line-height: 1.3;
    }}
    .app-topbar p {{ color: rgba(255,255,255,0.85); font-size: 0.95rem; margin: 0; max-width: 720px; line-height: 1.55; }}

    .kartu-utama {{
        background: {COLOR_SURFACE};
        border: 1px solid {COLOR_BORDER};
        padding: 28px 30px;
        border-radius: 16px;
        box-shadow: 0 3px 10px -4px rgba(22, 27, 51, 0.06);
        margin-bottom: 20px;
    }}
    .kartu-utama h3 {{ color: {COLOR_INK}; font-size: 1.15rem; margin: 0 0 6px 0; }}
    .kartu-utama .subjudul {{ color: {COLOR_MUTED}; font-size: 0.9rem; margin-bottom: 1rem; line-height: 1.55; }}

    div[data-testid="stVerticalBlockBorderWrapper"] {{
        background: {COLOR_SURFACE} !important;
        border: 1px solid {COLOR_BORDER} !important;
        border-radius: 16px !important;
        box-shadow: 0 3px 10px -4px rgba(22, 27, 51, 0.06);
    }}
    div[data-testid="stVerticalBlockBorderWrapper"] > div {{
        padding: 6px 4px;
    }}

    div[data-testid="stMetric"] {{
        background: {COLOR_SURFACE};
        border: 1px solid {COLOR_BORDER};
        padding: 16px 18px;
        border-radius: 14px;
        box-shadow: 0 2px 6px rgba(22,27,51,0.04);
    }}
    div[data-testid="stMetric"] label {{ color: {COLOR_MUTED} !important; font-weight: 700; text-transform: uppercase; font-size: 0.68rem; letter-spacing: 0.05em; }}
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {{ color: {COLOR_INK} !important; font-size: 1.45rem; font-weight: 800; }}

    .stButton > button, .stDownloadButton > button {{
        border-radius: 10px; font-weight: 700;
        background: linear-gradient(135deg, {COLOR_PRIMARY} 0%, {COLOR_PRIMARY_DK} 100%);
        color: white; border: none; padding: 0.7rem 1.4rem;
        box-shadow: 0 6px 14px -6px rgba(55, 48, 209, 0.5);
        transition: transform 0.15s ease, box-shadow 0.15s ease;
        width: 100%;
    }}
    .stButton > button:hover, .stDownloadButton > button:hover {{
        transform: translateY(-1px);
        box-shadow: 0 10px 18px -6px rgba(55, 48, 209, 0.55);
    }}

    .stButton > button[kind="secondary"] {{
        background: {COLOR_SURFACE} !important;
        color: {COLOR_PRIMARY} !important;
        border: 1.5px solid {COLOR_PRIMARY} !important;
        box-shadow: none !important;
    }}
    .stButton > button[kind="secondary"]:hover {{
        background: {COLOR_PRIMARY} !important;
        color: #ffffff !important;
        border-color: {COLOR_PRIMARY} !important;
        transform: translateY(-1px);
    }}
    .stButton > button[kind="secondary"]:hover * {{
        color: #ffffff !important;
    }}

    div[data-testid="stFileUploader"] {{
        background-color: #ffffff !important;
        border: 2px dashed #cbd5e1 !important;
        border-radius: 14px;
        padding: 16px;
    }}
    div[data-testid="stFileUploader"] section {{
        background-color: #ffffff !important;
        border: none !important;
        pointer-events: none !important;
        cursor: default !important;
    }}
    div[data-testid="stFileUploader"] section > input {{
        pointer-events: none !important;
    }}
    div[data-testid="stFileUploader"] section button,
    div[data-testid="stFileUploader"] section [data-testid="stBaseButton-secondary"] {{
        pointer-events: auto !important;
        cursor: pointer !important;
    }}
    div[data-testid="stFileUploader"] span,
    div[data-testid="stFileUploader"] label {{
        color: {COLOR_INK} !important;
    }}
    div[data-testid="stFileUploader"] section small,
    div[data-testid="stFileUploader"] section span:not(button span),
    div[data-testid="stFileUploader"] section p,
    div[data-testid="stFileUploader"] section label,
    div[data-testid="stFileUploader"] section > div,
    div[data-testid="stFileUploader"] section > div * {{
        font-size: 0.72rem !important;
        line-height: 1.4 !important;
        color: {COLOR_MUTED} !important;
    }}
    div[data-testid="stFileUploader"] section button,
    div[data-testid="stFileUploader"] section button * {{
        font-size: 0.9rem !important;
        color: #000000 !important;
        fill: #000000 !important;
    }}

    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"],
    div[data-testid="stFileUploader"] [data-testid="stUploadedFile"],
    div[data-testid="stFileUploader"] [data-baseweb="tag"] {{
        background-color: #ffffff !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 10px !important;
        box-shadow: 0 2px 6px rgba(22, 27, 51, 0.08) !important;
        padding: 8px 10px !important;
        pointer-events: auto !important;
    }}

    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] *,
    div[data-testid="stFileUploader"] [data-testid="stUploadedFile"] *,
    div[data-testid="stFileUploader"] [data-baseweb="tag"] *,
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFileName"] {{
        color: {COLOR_INK} !important;
        fill: {COLOR_INK} !important;
        font-weight: 700 !important;
        opacity: 1 !important;
        pointer-events: auto !important;
    }}

    div[data-testid="stFileUploader"] section button {{
        background-color: {COLOR_PRIMARY_BG} !important;
        color: {COLOR_PRIMARY} !important;
        border: 1px solid {COLOR_BORDER} !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        box-shadow: none !important;
    }}
    div[data-testid="stFileUploader"] section button:hover {{
        background-color: {COLOR_PRIMARY} !important;
        color: #ffffff !important;
        border-color: {COLOR_PRIMARY} !important;
    }}

    div.stAlert, div.stAlert p, div.stAlert span, div.stAlert li,
    div[data-testid="stNotification"], div[data-testid="stNotification"] p,
    div[data-testid="stAlertContainer"], div[data-testid="stAlertContainer"] p {{
        font-weight: 400 !important;
        font-size: 0.95rem !important;
    }}
    div.stAlert strong,
    div[data-testid="stAlertContainer"] strong,
    div[data-testid="stNotification"] strong {{
        font-weight: 400 !important;
    }}

    div.stSuccess, div.stSuccess p, div.stSuccess span, div.stSuccess li, div.stSuccess strong,
    div[data-testid="stAlertContentSuccess"], div[data-testid="stAlertContentSuccess"] *,
    div[data-testid="stNotificationContentSuccess"], div[data-testid="stNotificationContentSuccess"] * {{
        color: {COLOR_HIGH} !important;
    }}

    div.stInfo, div.stInfo p, div.stInfo span, div.stInfo li, div.stInfo strong,
    div[data-testid="stAlertContentInfo"], div[data-testid="stAlertContentInfo"] *,
    div[data-testid="stNotificationContentInfo"], div[data-testid="stNotificationContentInfo"] * {{
        color: {COLOR_INFO} !important;
    }}

    div.stWarning, div.stWarning p, div.stWarning span, div.stWarning li, div.stWarning strong,
    div[data-testid="stAlertContentWarning"], div[data-testid="stAlertContentWarning"] *,
    div[data-testid="stNotificationContentWarning"], div[data-testid="stNotificationContentWarning"] * {{
        color: {COLOR_MED} !important;
    }}

    div.stError, div.stError p, div.stError span, div.stError li, div.stError strong,
    div[data-testid="stAlertContentError"], div[data-testid="stAlertContentError"] *,
    div[data-testid="stNotificationContentError"], div[data-testid="stNotificationContentError"] * {{
        color: {COLOR_LOW} !important;
    }}

    div[data-testid="stRadio"] > div[role="radiogroup"] {{
        flex-direction: row !important;
        flex-wrap: wrap;
        gap: 10px !important;
    }}
    div[data-testid="stRadio"] label {{
        background-color: {COLOR_SURFACE} !important;
        border: 1px solid {COLOR_BORDER} !important;
        border-radius: 10px !important;
        padding: 10px 18px !important;
        cursor: pointer;
        transition: all 0.15s ease-in-out;
        margin: 0 !important;
    }}
    div[data-testid="stRadio"] label:hover {{
        background-color: {COLOR_PRIMARY_BG} !important;
        border-color: {COLOR_PRIMARY} !important;
    }}
    div[data-testid="stRadio"] label p,
    div[data-testid="stRadio"] label span,
    div[data-testid="stRadio"] label div {{
        color: {COLOR_MUTED} !important;
        font-weight: 700 !important;
        opacity: 1 !important;
        -webkit-text-fill-color: {COLOR_MUTED} !important;
    }}
    div[data-testid="stRadio"] label:has(input:checked) {{
        background-color: {COLOR_PRIMARY} !important;
        border-color: {COLOR_PRIMARY} !important;
    }}
    div[data-testid="stRadio"] label:has(input:checked) p,
    div[data-testid="stRadio"] label:has(input:checked) span,
    div[data-testid="stRadio"] label:has(input:checked) div {{
        color: #ffffff !important;
        font-weight: 800 !important;
        -webkit-text-fill-color: #ffffff !important;
    }}

    /* === PERBAIKAN: sembunyikan bulatan radio bawaan secara menyeluruh ===
       Sebelumnya hanya menyasar `label > div:first-child`, yang tidak selalu
       cocok dengan struktur DOM BaseWeb di breakpoint mobile. Sekarang
       menyasar beberapa kemungkinan elemen sekaligus supaya konsisten
       di desktop maupun HP. */
    div[data-testid="stRadio"] label > div:first-child,
    div[data-testid="stRadio"] label [data-baseweb="radio"],
    div[data-testid="stRadio"] label [data-baseweb="radio"] > div,
    div[data-testid="stRadio"] label [role="radio"] {{
        display: none !important;
        width: 0 !important;
        height: 0 !important;
        overflow: hidden !important;
        margin: 0 !important;
        padding: 0 !important;
    }}

    section[data-testid="stSidebar"] {{
        background-color: #ffffff;
        border-right: 1px solid {COLOR_BORDER};
        padding-top: 1rem;
    }}
    .kotak-judul-sidebar {{
        background: linear-gradient(135deg, {COLOR_PRIMARY_DK}, {COLOR_PRIMARY});
        padding: 20px 16px;
        border-radius: 14px;
        margin-bottom: 18px;
        text-align: center;
        box-shadow: 0 6px 14px -6px rgba(55,48,209,0.4);
        width: 100%;
        box-sizing: border-box;
    }}
    .kotak-judul-sidebar h3 {{ color: #ffffff; font-size: 1.05rem; margin: 0 0 3px 0; font-weight: 800; }}
    .kotak-judul-sidebar p {{ color: rgba(255,255,255,0.8); font-size: 0.68rem; margin: 0; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; }}

    .progres-sidebar {{
        margin-bottom: 18px;
    }}
    .progres-sidebar .progres-label {{
        display: flex;
        justify-content: space-between;
        font-size: 0.72rem;
        font-weight: 700;
        color: {COLOR_MUTED};
        text-transform: uppercase;
        letter-spacing: 0.04em;
        margin-bottom: 6px;
    }}
    .progres-sidebar .progres-label span:last-child {{ color: {COLOR_PRIMARY}; }}
    .progres-track {{
        width: 100%;
        height: 8px;
        background: {COLOR_BG};
        border-radius: 999px;
        overflow: hidden;
        border: 1px solid {COLOR_BORDER};
    }}
    .progres-fill {{
        height: 100%;
        background: linear-gradient(90deg, {COLOR_PRIMARY} 0%, #5b52e0 100%);
        border-radius: 999px;
        transition: width 0.3s ease;
    }}

    section[data-testid="stSidebar"] div.stRadio {{
        width: 100% !important;
    }}
    section[data-testid="stSidebar"] div.stRadio > div {{
        width: 100% !important;
    }}
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] {{
        flex-direction: column !important;
        gap: 12px;
        width: 100% !important;
    }}
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label {{
        background-color: {COLOR_BG} !important;
        padding: 12px 16px !important;
        border-radius: 12px !important;
        border: 1px solid {COLOR_BORDER} !important;
        transition: all 0.15s ease-in-out;
        width: 100% !important;
        box-sizing: border-box !important;
        margin: 0 !important;
        cursor: pointer;
    }}
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:hover {{
        background-color: {COLOR_PRIMARY_BG} !important;
        border-color: {COLOR_PRIMARY} !important;
    }}
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label p {{
        margin: 0 !important; padding: 0 !important; font-size: 0.87rem; font-weight: 600; color: #333a52 !important; line-height: normal !important;
        -webkit-text-fill-color: #333a52 !important;
    }}
    section[data-testid="stSidebar"] .stRadio input[type="radio"] {{
        accent-color: {COLOR_PRIMARY} !important;
    }}
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:has(input:checked) {{
        border-color: {COLOR_PRIMARY} !important;
        background-color: {COLOR_PRIMARY_BG} !important;
    }}

    hr {{ border-color: {COLOR_BORDER}; margin: 1.4rem 0; }}

    /* === PERBAIKAN: konsistensi tampilan sidebar khusus di layar HP ===
       Streamlit punya breakpoint mobile bawaan (<=640px) yang bisa membuat
       struktur/ukuran beberapa komponen berubah. Blok ini memaksa ulang
       styling utama supaya sidebar tetap konsisten dengan tampilan laptop. */
    @media (max-width: 640px) {{
        section[data-testid="stSidebar"] {{
            width: 100% !important;
            min-width: 100% !important;
        }}
        div[data-testid="stRadio"] label > div:first-child,
        div[data-testid="stRadio"] label [data-baseweb="radio"],
        div[data-testid="stRadio"] label [data-baseweb="radio"] > div,
        div[data-testid="stRadio"] label [role="radio"],
        div[data-testid="stRadio"] label span:first-child {{
            display: none !important;
            width: 0 !important;
            height: 0 !important;
        }}
        section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label {{
            padding: 12px 16px !important;
        }}
        section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label p {{
            font-size: 0.87rem !important;
        }}
        .kotak-judul-sidebar {{
            padding: 18px 14px;
        }}
        .app-topbar {{
            padding: 22px 20px;
        }}
        .app-topbar h1 {{
            font-size: 1.3rem;
        }}
    }}
    </style>
""", unsafe_allow_html=True)

# PATH & KONSTANTA MODEL

BASE_DIR = Path(__file__).resolve().parent
MODEL_PATH = BASE_DIR / "models" / "model_terbaik.json"
ENCODER_PATH = BASE_DIR / "models" / "label_encoder.pkl"
MODEL_FEATURES = ["Frekuensi", "Rata_Qty", "Rata_Harga"]

RECOMMENDATION_MAPPING = {
    "Potensi Tinggi": "Tingkatkan Stok",
    "Potensi Sedang": "Pertahankan Stok",
    "Potensi Rendah": "Kurangi Stok",
}

NAV_STEPS = [
    ("1. Upload Data", "📁", "Upload Data"),
    ("2. Pembersihan Data", "🧹", "Pembersihan Data"),
    ("3. Pembentukan Fitur", "⚙️", "Pembentukan Fitur"),
    ("4. Klasifikasi Potensi Penjualan", "🔍", "Klasifikasi Potensi Penjualan"),
    ("5. Dashboard & Visualisasi", "📊", "Dashboard & Visualisasi"),
    ("6. Unduh Laporan", "📥", "Unduh Laporan"),
]
NAV_KEYS = [step[0] for step in NAV_STEPS]

# FUNGSI CACHE / MODEL

@st.cache_resource
def load_trained_model():
    if not MODEL_PATH.exists():
        raise FileNotFoundError(
            f"File model tidak ditemukan di '{MODEL_PATH}'. "
            "Pastikan folder 'models/' ikut di-deploy bersama aplikasi ini."
        )
    if not ENCODER_PATH.exists():
        raise FileNotFoundError(
            f"File label encoder tidak ditemukan di '{ENCODER_PATH}'. "
            "Pastikan folder 'models/' ikut di-deploy bersama aplikasi ini."
        )

    model = xgb.XGBClassifier()
    model.load_model(str(MODEL_PATH))
    encoder = joblib.load(ENCODER_PATH)
    return model, encoder

COLUMN_KEYWORD_RULES = [
    ("harga", "Harga Satuan", ["harga satuan", "harga jual", "price", "satuan", "cost", "nilai", "harga"]),
    ("qty", "Qty", ["jumlah terjual", "qty", "jumlah", "quantity", "jml", "count", "banyak", "pcs", "volume", "terjual"]),
    ("tanggal", "Tanggal Faktur", ["tanggal faktur", "tgl transaksi", "tanggal", "tgl", "date", "faktur", "waktu", "hari"]),
    ("barang", "Nama Barang", ["nama barang", "nama produk", "barang", "item", "menu", "product", "produk"]),
]

def _build_column_mapping(columns):
    col_mapping = {}
    mapped_check = set()
    ambiguous = []

    for raw_col in columns:
        low_col = str(raw_col).strip().lower()
        matched_this_col = False

        for key, target_name, keywords in COLUMN_KEYWORD_RULES:
            if any(kw in low_col for kw in keywords):
                matched_this_col = True
                if key not in mapped_check:
                    col_mapping[raw_col] = target_name
                    mapped_check.add(key)
                else:
                    ambiguous.append((raw_col, target_name))
                break

        if not matched_this_col:
            continue

    return col_mapping, ambiguous

def auto_detect_columns(raw_df):
    clean_df = raw_df.loc[:, ~raw_df.columns.duplicated()].copy()
    col_mapping, _ = _build_column_mapping(clean_df.columns)
    clean_df = clean_df.rename(columns=col_mapping)
    return clean_df

def detect_column_mapping_issues(raw_df):
    clean_cols = raw_df.loc[:, ~raw_df.columns.duplicated()].columns
    _, ambiguous = _build_column_mapping(clean_cols)
    return ambiguous

@st.cache_data
def clean_transaction_data(input_df):
    processed_df = auto_detect_columns(input_df)
    required_cols = ["Tanggal Faktur", "Nama Barang", "Qty", "Harga Satuan"]

    for col in required_cols:
        if col not in processed_df.columns:
            raise ValueError(f"Kolom wajib '{col}' tidak ditemukan. Cek lagi header excelnya ya.")

    processed_df = processed_df.loc[:, ~processed_df.columns.duplicated()].copy()

    duplicates_qty = int(processed_df.duplicated().sum())
    if duplicates_qty > 0:
        processed_df = processed_df.drop_duplicates()

    processed_df["Nama Barang"] = processed_df["Nama Barang"].astype(str).str.strip()
    processed_df["Tanggal Faktur"] = pd.to_datetime(processed_df["Tanggal Faktur"], errors="coerce")
    processed_df["Qty"] = pd.to_numeric(processed_df["Qty"], errors="coerce")
    processed_df["Harga Satuan"] = pd.to_numeric(processed_df["Harga Satuan"], errors="coerce")

    missing_qty = int(processed_df.isnull().sum().sum())
    processed_df = processed_df.dropna(subset=required_cols)
    processed_df = processed_df[(processed_df["Qty"] > 0) & (processed_df["Harga Satuan"] > 0)]

    summary_info = {
        "duplikat_dihapus": duplicates_qty,
        "missing_value": missing_qty,
        "total_bersih": len(processed_df)
    }
    return processed_df, summary_info

@st.cache_data
def extract_product_features(clean_df):
    features_df = clean_df.groupby("Nama Barang").agg(
        Total_Qty=("Qty", "sum"),
        Rata_Qty=("Qty", "mean"),
        Frekuensi=("Qty", "count"),
        Rata_Harga=("Harga Satuan", "mean")
    ).reset_index()

    features_df["Rata_Qty"] = features_df["Rata_Qty"].round(2)
    features_df["Rata_Harga"] = features_df["Rata_Harga"].round(2)

    return features_df

def run_model_prediction(model, encoder, df):
    return pd.Series(
        encoder.inverse_transform(model.predict(df[MODEL_FEATURES]).astype(int)),
        index=df.index
    )

def apply_plotly_theme(fig, title_text):
    fig.update_layout(
        template=PLOTLY_TEMPLATE,
        font=PLOTLY_FONT,
        title=dict(text=title_text, font=dict(size=16, family="Poppins, sans-serif", color=COLOR_INK)),
        legend=dict(orientation="h", yanchor="bottom", y=1.04, xanchor="right", x=1),
        margin=dict(t=70, l=10, r=30, b=10),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(tickfont=dict(color=COLOR_INK, size=12), title_font=dict(color=COLOR_INK)),
        yaxis=dict(tickfont=dict(color=COLOR_INK, size=12), title_font=dict(color=COLOR_INK)),
    )
    return fig

def render_topbar(title, desc):
    st.markdown(f"""
        <div class="app-topbar">
            <div class="eyebrow" style="color: rgba(255,255,255,0.75); font-size: 0.72rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.14em; margin-bottom: 6px;">Sistem Pendukung Keputusan · UMKM</div>
            <h1>{title}</h1>
            <p>{desc}</p>
        </div>
    """, unsafe_allow_html=True)

def render_card_open(title, subtitle=""):
    sub_html = f'<p class="subjudul">{subtitle}</p>' if subtitle else ""
    st.markdown(f"""
        <div class="kartu-utama">
            <h3>{title}</h3>
            {sub_html}
    """, unsafe_allow_html=True)

def render_card_close():
    st.markdown("</div>", unsafe_allow_html=True)

def go_to_step(step_key):
    st.session_state["nav_active"] = step_key
    st.rerun()

def render_next_step_button(current_index, label_override=None):
    if current_index >= len(NAV_STEPS) - 1:
        return
    next_key = NAV_KEYS[current_index + 1]
    next_icon = NAV_STEPS[current_index + 1][1]
    next_label = NAV_STEPS[current_index + 1][2]
    button_text = label_override or f"Lanjut ke {next_icon} {next_label} →"
    st.markdown("<div style='margin-top: 6px;'></div>", unsafe_allow_html=True)
    col_spacer, col_btn = st.columns([2, 1.3])
    with col_btn:
        if st.button(button_text, key=f"lanjut_dari_{current_index}", use_container_width=True, type="secondary"):
            go_to_step(next_key)

# INISIALISASI SESSION STATE & NAVIGASI YANG STABIL

for state_name in ["data_mentah", "data_bersih", "info_bersih", "data_fitur", "data_hasil", "nama_file_aktif"]:
    if state_name not in st.session_state:
        st.session_state[state_name] = None

if "nav_active" not in st.session_state:
    st.session_state["nav_active"] = NAV_KEYS[0]

if st.session_state.get("_pending_nav"):
    st.session_state["nav_active"] = st.session_state["_pending_nav"]
    st.session_state["_pending_nav"] = None

if st.session_state.get("nav_active") not in NAV_KEYS:
    st.session_state["nav_active"] = NAV_KEYS[0]

STEP_COMPLETION = [
    st.session_state["data_mentah"] is not None,
    st.session_state["data_bersih"] is not None,
    st.session_state["data_fitur"] is not None,
    st.session_state["data_hasil"] is not None,
    st.session_state["data_hasil"] is not None,
    st.session_state["data_hasil"] is not None,
]
PIPELINE_DONE_COUNT = sum(STEP_COMPLETION[:4])

# SIDEBAR: JUDUL + PROGRES + NAVIGASI

st.sidebar.markdown("""
    <div class="kotak-judul-sidebar">
        <h3>📌 Tahapan Analisis</h3>
        <p>Sistem Klasifikasi Stok</p>
    </div>
""", unsafe_allow_html=True)

progres_persen = int((PIPELINE_DONE_COUNT / 4) * 100)
st.sidebar.markdown(f"""
    <div class="progres-sidebar">
        <div class="progres-label">
            <span>Progres Tahap Inti</span>
            <span>{PIPELINE_DONE_COUNT}/4</span>
        </div>
        <div class="progres-track">
            <div class="progres-fill" style="width: {progres_persen}%;"></div>
        </div>
    </div>
""", unsafe_allow_html=True)

def _format_nav_label(step_key):
    idx = NAV_KEYS.index(step_key)
    original_icon = NAV_STEPS[idx][1]
    label_text = step_key[3:]
    icon = "✅" if STEP_COMPLETION[idx] else original_icon
    return f"{icon}  {label_text}"

current_nav_index = NAV_KEYS.index(st.session_state["nav_active"]) if st.session_state["nav_active"] in NAV_KEYS else 0

nav = st.sidebar.radio(
    "Navigasi:",
    NAV_KEYS,
    index=current_nav_index,
    format_func=_format_nav_label,
    label_visibility="collapsed"
)

if nav != st.session_state["nav_active"]:
    st.session_state["nav_active"] = nav
    st.rerun()

st.sidebar.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)

if st.sidebar.button("🔄 Mulai Ulang Analisis", use_container_width=True):
    for state_name in ["data_mentah", "data_bersih", "info_bersih", "data_fitur",
                       "data_hasil", "nama_file_aktif", "_kolom_ambigu_terakhir"]:
        st.session_state[state_name] = None
    st.session_state["nav_active"] = NAV_KEYS[0]
    st.rerun()

# MENU 1: UPLOAD DATA

if nav == "1. Upload Data":
    render_topbar(
        "Sistem Pendukung Keputusan Analisis Potensi Penjualan Produk UMKM",
        "Aplikasi cerdas ini membantu Anda mengelompokkan produk berdasarkan potensi penjualannya sekaligus memberikan rekomendasi pengelolaan stok yang tepat berdasarkan riwayat transaksi toko Anda."
    )

    render_card_open(
        "📁 Unggah Laporan Data Penjualan",
        "Silahkan unggah file laporan penjualan Anda dalam format <b>Excel (.xlsx)</b>. Sistem akan secara otomatis membaca data penting seperti <b>Tanggal Transaksi</b>, <b>Nama Produk</b>, <b>Jumlah Terjual (Qty)</b>, dan <b>Harga Satuan</b> tanpa perlu pengaturan rumit."
    )

    uploaded_file = st.file_uploader("Pilih file Excel laporan data penjualan (.xlsx)", type=["xlsx"], key="uploader_excel")

    if uploaded_file is not None:
        try:
            if st.session_state["nama_file_aktif"] != uploaded_file.name:
                with st.spinner("Membaca dan mendeteksi struktur kolom..."):
                    raw_df = pd.read_excel(uploaded_file)
                    raw_df.columns = raw_df.columns.str.strip()

                    ambiguous_cols = detect_column_mapping_issues(raw_df)
                    detected_df = auto_detect_columns(raw_df)

                st.session_state["data_mentah"] = detected_df
                st.session_state["nama_file_aktif"] = uploaded_file.name
                st.session_state["data_bersih"] = None
                st.session_state["info_bersih"] = None
                st.session_state["data_fitur"] = None
                st.session_state["data_hasil"] = None

                if ambiguous_cols:
                    daftar_kolom = ", ".join(f"'{c}' → {t}" for c, t in ambiguous_cols)
                    st.session_state["_kolom_ambigu_terakhir"] = daftar_kolom
                else:
                    st.session_state["_kolom_ambigu_terakhir"] = None
                st.rerun()
        except Exception as e:
            st.error(f"Gagal membaca file excel: {e}")

    render_card_close()

    if st.session_state.get("_kolom_ambigu_terakhir"):
        st.warning(
            "Beberapa kolom terdeteksi cocok dengan kategori yang sudah terisi, "
            f"sehingga diabaikan oleh sistem: {st.session_state['_kolom_ambigu_terakhir']}. "
            "Jika ini seharusnya kolom penting, ganti nama header di file Excel Anda agar lebih spesifik."
        )

    if st.session_state["data_mentah"] is not None:
        df_preview = st.session_state["data_mentah"]

        item_col = df_preview["Nama Barang"] if "Nama Barang" in df_preview.columns else df_preview.iloc[:, 1]
        if isinstance(item_col, pd.DataFrame):
            item_col = item_col.iloc[:, 0]
        total_unique_items = int(item_col.nunique())

        qty_col = df_preview["Qty"] if "Qty" in df_preview.columns else df_preview.iloc[:, 2]
        if isinstance(qty_col, pd.DataFrame):
            qty_col = qty_col.iloc[:, 0]
        sum_qty_all = int(pd.to_numeric(qty_col, errors='coerce').sum())

        date_col = df_preview["Tanggal Faktur"] if "Tanggal Faktur" in df_preview.columns else df_preview.iloc[:, 0]
        if isinstance(date_col, pd.DataFrame):
            date_col = date_col.iloc[:, 0]
        valid_dates = pd.to_datetime(date_col, errors="coerce").dropna()
        date_range_str = f"{valid_dates.min():%d/%m/%Y} - {valid_dates.max():%d/%m/%Y}" if not valid_dates.empty else "Tidak terdeteksi"

        st.success(f"Berhasil diunggah: {st.session_state['nama_file_aktif']} siap diproses.")

        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("TOTAL BARIS", f"{len(df_preview):,}")
        col2.metric("JUMLAH PRODUK", f"{total_unique_items:,}")
        col3.metric("TOTAL QTY", f"{sum_qty_all:,}")
        col4.metric("PERIODE", date_range_str)

        st.markdown("<br>", unsafe_allow_html=True)
        with st.expander("🔎 Tampilkan Pratinjau 10 Baris Pertama Data"):
            st.dataframe(df_preview.head(10), use_container_width=True, hide_index=True)

        render_next_step_button(0)
    else:
        st.info("Silakan unggah file laporan data penjualan terlebih dahulu untuk memulai analisis.")

# MENU 2: PEMBERSIHAN DATA

elif nav == "2. Pembersihan Data":
    render_card_open(
        "🧹 Pembersihan Data (Data Cleaning)",
        "Menu ini berfungsi untuk merapihkan dan memastikan data penjualan Anda bersih dari kesalahan. "
        "Sistem secara otomatis menghapus data ganda, membuang data yang kosong, serta menyaring transaksi "
        "yang memiliki jumlah dan harga jual valid (di atas nol) agar siap diproses ke tahap berikutnya."
    )

    if st.session_state["data_mentah"] is not None:
        st.markdown("<div style='margin-top: 4px;'></div>", unsafe_allow_html=True)
        if st.button("✨ Jalankan Proses Pembersihan", type="primary"):
            try:
                with st.spinner("Membersihkan data transaksi..."):
                    clean_df, info_dict = clean_transaction_data(st.session_state["data_mentah"])
                st.session_state["data_bersih"] = clean_df
                st.session_state["info_bersih"] = info_dict
                if info_dict["total_bersih"] == 0:
                    st.warning(
                        "Setelah dibersihkan, tidak ada baris data yang tersisa. "
                        "Cek kembali isi kolom Qty/Harga Satuan pada file Anda (harus berupa angka > 0)."
                    )
                else:
                    st.success("Pembersihan data berhasil dilakukan!")
            except ValueError as ve:
                st.error(f"Gagal membersihkan data: {ve}")

        render_card_close()

        if st.session_state["data_bersih"] is not None:
            info_tampil = st.session_state["info_bersih"]
            st.markdown("<br>", unsafe_allow_html=True)
            col1, col2, col3 = st.columns(3)
            col1.metric("DATA DUPLIKAT", f"{info_tampil['duplikat_dihapus']} baris")
            col2.metric("DATA KOSONG", f"{info_tampil['missing_value']} anomali")
            col3.metric("DATA BERSIH", f"{info_tampil['total_bersih']} baris")

            st.markdown("<br>", unsafe_allow_html=True)
            st.info(
                f"Data Anda sudah berhasil dirapikan! Ada {info_tampil['duplikat_dihapus']} baris data ganda "
                f"dan {info_tampil['missing_value']} data kosong yang otomatis dibersihkan. "
                f"Kini tersisa total {info_tampil['total_bersih']} baris data bersih yang siap dilanjutkan "
                f"ke tahap berikutnya."
            )

            st.markdown("<br>", unsafe_allow_html=True)
            with st.expander("🔎 Tampilkan Sampel Data yang Sudah Bersih"):
                st.dataframe(st.session_state["data_bersih"].head(10), use_container_width=True, hide_index=True)

            if info_tampil["total_bersih"] > 0:
                render_next_step_button(1)
    else:
        render_card_close()
        st.warning("Belum ada data. Selesaikan Menu 1 (Upload Data) terlebih dahulu!")

# MENU 3: PEMBENTUKAN FITUR

elif nav == "3. Pembentukan Fitur":
    render_card_open(
        "⚙️ Pembentukan Fitur (Feature Engineering)",
        "Menu ini digunakan untuk merangkum seluruh data transaksi penjualan yang sebelumnya terpisah-pisah menjadi satu ringkasan data per produk. Hasil ringkasan inilah yang nantinya akan dibaca oleh sistem kecerdasan buatan untuk melakukan klasifikasi.<br><br>"
        "<b>Keterangan Kolom pada Tabel Ringkasan:</b><br>"
        "• <b>Nama Barang:</b> Nama produk atau menu yang terjual.<br>"
        "• <b>Total_Qty:</b> Total keseluruhan jumlah barang yang terjual dari awal hingga akhir periode.<br>"
        "• <b>Rata_Qty:</b> Rata-rata jumlah barang yang biasanya terjual dalam sekali transaksi.<br>"
        "• <b>Frekuensi:</b> Seberapa sering produk tersebut dibeli atau muncul dalam catatan transaksi.<br>"
        "• <b>Rata_Harga:</b> Rata-rata harga jual produk tersebut berdasarkan catatan transaksi."
    )

    if st.session_state["data_bersih"] is not None:
        if st.button("⚡ Mulai Pembentukan Fitur", type="primary"):
            with st.spinner("Merangkum transaksi menjadi fitur per produk..."):
                feat_df = extract_product_features(st.session_state["data_bersih"])
            st.session_state["data_fitur"] = feat_df
            st.success("Berhasil! Seluruh data transaksi berhasil dirangkum menjadi ringkasan produk yang siap dianalisis.")
    else:
        st.warning("Selesaikan Menu 2 (Pembersihan Data) terlebih dahulu!")

    render_card_close()

    if st.session_state["data_fitur"] is not None:
        df_tampil_fitur = st.session_state["data_fitur"]
        st.markdown("<br>", unsafe_allow_html=True)
        st.metric("TOTAL JENIS PRODUK", f"{len(df_tampil_fitur)} Produk")

        st.markdown("<br>", unsafe_allow_html=True)
        st.write("**Tabel Ringkasan Fitur Produk:**")
        st.dataframe(df_tampil_fitur, use_container_width=True, hide_index=True)

        render_next_step_button(2)

# MENU 4: KLASIFIKASI POTENSI PENJUALAN

elif nav == "4. Klasifikasi Potensi Penjualan":
    render_card_open(
        "🔍 Klasifikasi Potensi Penjualan",
        "Menu ini menggunakan sistem cerdas untuk mengelompokkan setiap produk ke dalam kategori potensi penjualan "
        "(Tinggi, Sedang, atau Rendah) berdasarkan data transaksi sebelumnya. Sistem secara otomatis memberikan "
        "rekomendasi pengelolaan stok: tingkatkan stok untuk produk berpotensi tinggi, pertahankan untuk potensi sedang, "
        "dan kurangi untuk potensi rendah agar modal usaha Anda lebih optimal."
    )
    render_card_close()

    if st.session_state["data_fitur"] is not None:
        try:
            with st.spinner("Memuat model dan menjalankan prediksi..."):
                model_obj, encoder_obj = load_trained_model()
                df_temp = st.session_state["data_fitur"].copy()
                df_temp["Kategori Potensi"] = run_model_prediction(model_obj, encoder_obj, df_temp)
                df_temp["Rekomendasi Stok"] = df_temp["Kategori Potensi"].map(RECOMMENDATION_MAPPING)
            st.session_state["data_hasil"] = df_temp

            st.success("Proses klasifikasi berhasil dijalankan!")
            st.markdown("<br>", unsafe_allow_html=True)

            c1, c2 = st.columns(2)
            with c1:
                keyword = st.text_input("🔎 Cari Berdasarkan Nama Produk", placeholder="Ketik nama produk...")
            with c2:
                selected_cat = st.multiselect(
                    "Filter Kategori Potensi",
                    options=["Potensi Tinggi", "Potensi Sedang", "Potensi Rendah"]
                )

            filtered_result = df_temp.copy()
            if keyword:
                filtered_result = filtered_result[filtered_result["Nama Barang"].str.contains(keyword, case=False, na=False)]
            if selected_cat:
                filtered_result = filtered_result[filtered_result["Kategori Potensi"].isin(selected_cat)]

            st.markdown("<br>", unsafe_allow_html=True)
            st.caption(f"Menampilkan {len(filtered_result)} dari {len(df_temp)} produk.")
            st.dataframe(
                filtered_result,
                use_container_width=True,
                hide_index=True
            )

            render_next_step_button(3)
        except Exception as err_msg:
            st.error(f"Terjadi kendala saat eksekusi model ML: {err_msg}")
    else:
        st.warning("Selesaikan Menu 3 (Pembentukan Fitur) terlebih dahulu.")

# MENU 5: DASHBOARD & VISUALISASI

elif nav == "5. Dashboard & Visualisasi":
    render_card_open(
        "📊 Dashboard & Visualisasi Grafik",
        "Menu ini menampilkan grafik yang interaktif dan mudah dipahami. Anda bisa melihat perbandingan "
        "kategori produk, produk apa saja yang paling laris, serta bagaimana perkembangan tren penjualan "
        "toko Anda dari waktu ke waktu."
    )
    render_card_close()

    if st.session_state["data_hasil"] is not None:
        df_dash = st.session_state["data_hasil"]

        count_high = int((df_dash["Kategori Potensi"] == "Potensi Tinggi").sum())
        count_med = int((df_dash["Kategori Potensi"] == "Potensi Sedang").sum())
        count_low = int((df_dash["Kategori Potensi"] == "Potensi Rendah").sum())
        total_items_all = len(df_dash)

        col1, col2, col3 = st.columns(3)
        col1.metric("🟢 POTENSI TINGGI", f"{count_high} Produk")
        col2.metric("🟠 POTENSI SEDANG", f"{count_med} Produk")
        col3.metric("🔴 POTENSI RENDAH", f"{count_low} Produk")

        st.markdown("---")

        pie_df = df_dash["Kategori Potensi"].value_counts().reset_index()
        pie_df.columns = ["Kategori", "Jumlah Produk"]

        pilihan_visual = st.radio(
            "Pilih Visualisasi:",
            ["📌 Proporsi Kategori", "🏆 10 Produk Terbesar", "📈 Tren Historis Bulanan"],
            horizontal=True,
            label_visibility="collapsed",
            key="pilihan_dashboard_visual"
        )
        st.markdown("<div style='margin-top: 4px;'></div>", unsafe_allow_html=True)

        if pilihan_visual == "📌 Proporsi Kategori":
            fig_pie = px.pie(
                pie_df, names="Kategori", values="Jumlah Produk",
                color="Kategori", color_discrete_map=CHART_COLOR_MAP, hole=0.55
            )
            fig_pie.update_traces(textinfo="percent+label", textfont_size=12)
            apply_plotly_theme(fig_pie, "Distribusi Klasifikasi Potensi Penjualan Produk")
            fig_pie.update_layout(height=360)
            st.plotly_chart(fig_pie, use_container_width=True)

            pie_sorted = pie_df.sort_values("Jumlah Produk", ascending=False).reset_index(drop=True)
            top_kategori = pie_sorted.loc[0, "Kategori"]
            top_kategori_pct = pie_sorted.loc[0, "Jumlah Produk"] / total_items_all * 100
            st.info(
                f"💡 **Saran untuk Toko Anda:** Dari total {total_items_all} produk yang dianalisis, sebagian besar produk "
                f"ternyata masuk dalam kategori **{top_kategori}** ({top_kategori_pct:.1f}%). "
                f"Supaya modal usaha Anda berputar dengan lancar, buat paket promo atau bundling untuk produk yang lambat terjual agar cepat habis. "
                f"Selain itu, pastikan Anda selalu mengamankan stok dan memperbanyak persediaan untuk produk yang berpotensi tinggi "
                f"supaya pelanggan tidak kecewa kehabisan barang."
            )

        elif pilihan_visual == "🏆 10 Produk Terbesar":
            top_10 = df_dash.sort_values(by="Total_Qty", ascending=False).head(10)
            max_qty_val = top_10["Total_Qty"].max()
            fig_bar = px.bar(
                top_10, x="Total_Qty", y="Nama Barang", orientation='h',
                color="Kategori Potensi", color_discrete_map=CHART_COLOR_MAP,
                text="Total_Qty"
            )
            fig_bar.update_traces(
                texttemplate='%{text:,.0f}',
                textposition='outside',
                textfont=dict(color=COLOR_INK, size=12),
                cliponaxis=False
            )
            fig_bar.update_layout(
                yaxis={'categoryorder': 'total ascending', 'tickfont': dict(color=COLOR_INK, size=12)},
                xaxis={'range': [0, max_qty_val * 1.18]},
                xaxis_title="Total Qty Terjual",
                yaxis_title=""
            )
            apply_plotly_theme(fig_bar, "10 Produk dengan Akumulasi Penjualan Tertinggi")
            fig_bar.update_layout(height=420, margin=dict(l=10, r=60, t=70, b=10))
            st.plotly_chart(fig_bar, use_container_width=True)

            top_produk_nama = top_10.iloc[0]["Nama Barang"]
            top_produk_qty = int(top_10.iloc[0]["Total_Qty"])
            top_produk_kategori = top_10.iloc[0]["Kategori Potensi"]
            st.info(
                f"💡 **Saran untuk Toko Anda:** Wah, produk **{top_produk_nama}** jadi juara di toko Anda dengan total penjualan mencapai "
                f"{top_produk_qty:,} unit dan berstatus **{top_produk_kategori}**! "
                f"Karena produk ini paling diandalkan pembeli, jaga terus ketersediaan stoknya dan jadikan sorotan utama "
                f"untuk promosi agar omzet toko Anda semakin meningkat."
            )

        else:
            df_trend = st.session_state["data_bersih"].copy()
            df_trend["Bulan"] = df_trend["Tanggal Faktur"].dt.to_period("M").dt.to_timestamp()
            monthly_rekap = df_trend.groupby("Bulan").agg(Total_Qty=("Qty", "sum")).reset_index().sort_values("Bulan")

            monthly_rekap["Bulan_Label"] = monthly_rekap["Bulan"].dt.strftime("%b %Y")

            fig_line = px.area(
                monthly_rekap,
                x="Bulan_Label",
                y="Total_Qty",
                markers=True,
                labels={"Bulan_Label": "Bulan"}
            )
            fig_line.update_traces(line_color=COLOR_PRIMARY, fillcolor="rgba(55,48,209,0.12)")
            apply_plotly_theme(fig_line, "Grafik Tren Historis Kuantitas Penjualan Bulanan")
            fig_line.update_layout(
                height=360,
                xaxis=dict(
                    title="Bulan",
                    tickfont=dict(color=COLOR_INK, size=12),
                    title_font=dict(color=COLOR_INK),
                    categoryorder="array",
                    categoryarray=monthly_rekap["Bulan_Label"].tolist()
                )
            )
            st.plotly_chart(fig_line, use_container_width=True)

            if len(monthly_rekap) >= 2:
                qty_awal = monthly_rekap.iloc[0]["Total_Qty"]
                qty_akhir = monthly_rekap.iloc[-1]["Total_Qty"]
                arah_tren = "meningkat" if qty_akhir >= qty_awal else "menurun"
                persen_perubahan = abs((qty_akhir - qty_awal) / qty_awal * 100) if qty_awal > 0 else 0
                st.info(
                    f"💡 **Saran untuk Toko Anda:** Dari catatan bulanan, tren penjualan toko Anda terlihat **{arah_tren}** sekitar {persen_perubahan:.1f}% "
                    f"dari {monthly_rekap.iloc[0]['Bulan']:%b %Y} ({int(qty_awal):,} unit) ke {monthly_rekap.iloc[-1]['Bulan']:%b %Y} ({int(qty_akhir):,} unit). "
                    f"Kalau penjualannya sedang naik, yuk siap-siap tambah stok dari awal supaya tidak keteteran melayani pembeli. "
                    f"Sebaliknya, kalau trennya melambat, kurangi belanja barang yang kurang laku agar kas toko tetap aman."
                )
            else:
                st.info("💡 **Saran untuk Toko Anda:** Catatan transaksi historis pada data yang diunggah belum mencakup rentang waktu lebih dari satu bulan, sehingga grafik tren bulanan belum dapat menampilkan perbandingan antar bulan.")

        render_next_step_button(4, label_override="Lanjut ke 📥 Unduh Laporan →")
    else:
        st.warning("Jalankan Menu 4 (Prediksi & Klasifikasi) terlebih dahulu.")

# MENU 6: UNDUH LAPORAN

elif nav == "6. Unduh Laporan":
    render_card_open(
        "📥 Unduh Laporan Hasil Analisis",
        "Di menu ini, Anda bisa mengunduh seluruh hasil analisis dan rekomendasi pengelolaan stok produk "
        "dalam bentuk file Excel (.xlsx) agar mudah disimpan, dibagikan, atau dicetak."
    )

    if st.session_state["data_hasil"] is not None:
        tabel_export = st.session_state["data_hasil"]

        KATEGORI_FILL_MAP = {
            "Potensi Tinggi": PatternFill(start_color=hex_no_hash(COLOR_HIGH_BG), end_color=hex_no_hash(COLOR_HIGH_BG), fill_type="solid"),
            "Potensi Sedang": PatternFill(start_color=hex_no_hash(COLOR_MED_BG), end_color=hex_no_hash(COLOR_MED_BG), fill_type="solid"),
            "Potensi Rendah": PatternFill(start_color=hex_no_hash(COLOR_LOW_BG), end_color=hex_no_hash(COLOR_LOW_BG), fill_type="solid"),
        }
        KATEGORI_FONT_MAP = {
            "Potensi Tinggi": Font(name="Calibri", size=11, bold=True, color=hex_no_hash(COLOR_HIGH)),
            "Potensi Sedang": Font(name="Calibri", size=11, bold=True, color=hex_no_hash(COLOR_MED)),
            "Potensi Rendah": Font(name="Calibri", size=11, bold=True, color=hex_no_hash(COLOR_LOW)),
        }

        excel_buffer = io.BytesIO()
        with st.spinner("Menyusun dan memformat file Excel..."), \
             pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            tabel_export.to_excel(writer, index=False, sheet_name="Hasil Klasifikasi")

            wb_file = writer.book
            ws_file = writer.sheets["Hasil Klasifikasi"]

            header_fill = PatternFill(start_color="161B33", end_color="161B33", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            box_border = Border(
                left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
            )

            col_names = list(tabel_export.columns)
            kategori_col_idx = col_names.index("Kategori Potensi") + 1 if "Kategori Potensi" in col_names else None
            nama_col_idx = col_names.index("Nama Barang") + 1 if "Nama Barang" in col_names else 1
            center_cols = {col_names.index(c) + 1 for c in ["Kategori Potensi", "Rekomendasi Stok"] if c in col_names}

            for c_idx in range(1, len(col_names) + 1):
                cell_h = ws_file.cell(row=1, column=c_idx)
                cell_h.fill = header_fill
                cell_h.font = header_font
                cell_h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell_h.border = box_border

            for r_idx in range(2, len(tabel_export) + 2):
                kategori_val = ws_file.cell(row=r_idx, column=kategori_col_idx).value if kategori_col_idx else None
                for c_idx in range(1, len(col_names) + 1):
                    cell_d = ws_file.cell(row=r_idx, column=c_idx)
                    cell_d.font = Font(name="Calibri", size=11)
                    cell_d.border = box_border

                    col_key = col_names[c_idx - 1]
                    if col_key == "Rata_Harga":
                        cell_d.number_format = '#,##0.00'
                        cell_d.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_key == "Rata_Qty":
                        cell_d.number_format = '0.00'
                        cell_d.alignment = Alignment(horizontal="right", vertical="center")
                    elif c_idx == nama_col_idx:
                        cell_d.alignment = Alignment(horizontal="left", vertical="center")
                    elif c_idx in center_cols:
                        cell_d.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell_d.alignment = Alignment(horizontal="right", vertical="center")

                    if kategori_val in KATEGORI_FILL_MAP and c_idx in center_cols:
                        cell_d.fill = KATEGORI_FILL_MAP[kategori_val]
                        cell_d.font = KATEGORI_FONT_MAP[kategori_val]

            ws_file.freeze_panes = "A2"
            for column_cells in ws_file.columns:
                max_width_val = 0
                col_letter_str = get_column_letter(column_cells[0].column)
                for cell_item in column_cells:
                    try:
                        if cell_item.value:
                            max_width_val = max(max_width_val, len(str(cell_item.value)))
                    except Exception:
                        pass
                ws_file.column_dimensions[col_letter_str].width = max(max_width_val + 5, 16)

        excel_buffer.seek(0)
        st.success("File laporan Excel Anda sudah siap untuk diunduh!")
        st.markdown("<br>", unsafe_allow_html=True)
        st.download_button(
            label="⬇️ Unduh Dokumen Laporan Rekomendasi (.xlsx)",
            data=excel_buffer,
            file_name="Laporan_Analisis_Potensi_Penjualan_UMKM.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.warning("Belum terdapat data hasil analisis. Harap selesaikan proses hingga Menu 4 terlebih dahulu.")

    render_card_close()